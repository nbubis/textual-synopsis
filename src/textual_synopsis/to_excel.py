import glob
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, Border, Side


def load_aligned_texts(directory="."):
    files = sorted(glob.glob(os.path.join(directory, "aligned_*.txt")))
    texts = []

    for f in files:
        with open(f, "r", encoding="utf-8") as f_obj:
            content = f_obj.read()
            # The filenames are aligned_XYZ.txt. Clean name for row label
            name = os.path.basename(f).replace("aligned_", "").replace(".txt", "")
            texts.append({"name": name, "content": content})

    return texts


def align_to_words(texts):
    if not texts:
        return []

    rows = []
    gap_char = "@"

    for t in texts:
        content = t["content"].strip()
        if not content:
            words = []
        else:
            words = content.split(" ")
        
        clean_words = ["" if w == gap_char else w for w in words]
        rows.append(clean_words)

    if not rows:
        return []

    length = len(rows[0])
    for i, r in enumerate(rows):
        if len(r) != length:
            raise ValueError(
                f"Length mismatch: {texts[i]['name']} has {len(r)} words vs {length} words"
            )

    # Filter out columns where all rows are empty strings
    valid_cols = []
    for col_idx in range(length):
        if any(rows[row_idx][col_idx] != "" for row_idx in range(len(rows))):
            valid_cols.append(col_idx)

    filtered_rows = [[] for _ in rows]
    for col_idx in valid_cols:
        for row_idx in range(len(rows)):
            filtered_rows[row_idx].append(rows[row_idx][col_idx])

    return filtered_rows


def create_printable_chunks(df, chunk_size=20):
    """
    Create a printable version of the DataFrame by chunking columns.

    Args:
        df: DataFrame with sources as rows and words as columns
        chunk_size: Number of word columns per chunk (default 20 for A4 landscape)

    Returns:
        DataFrame with chunks stacked vertically, separated by blank rows
    """
    num_cols = len(df.columns)
    max_cols = chunk_size  # Maximum columns in any chunk
    chunks = []

    # Split into chunks, padding each to the same width
    for start_col in range(0, num_cols, chunk_size):
        end_col = min(start_col + chunk_size, num_cols)
        chunk_df = df.iloc[:, start_col:end_col].copy()

        # Pad chunk with empty columns to reach chunk_size
        cols_to_add = max_cols - len(chunk_df.columns)
        if cols_to_add > 0:
            for j in range(cols_to_add):
                chunk_df[len(chunk_df.columns)] = ""

        # Reset column names to 0, 1, 2, ... for consistent alignment
        chunk_df.columns = range(len(chunk_df.columns))
        chunks.append(chunk_df)

    # Stack chunks vertically with blank rows between them
    result_chunks = []
    for i, chunk in enumerate(chunks):
        result_chunks.append(chunk)
        if i < len(chunks) - 1:  # Add blank row between chunks (except after last)
            # Create a blank row with empty strings
            blank_row = pd.DataFrame(
                [[""] * max_cols], columns=range(max_cols), index=[""]
            )
            result_chunks.append(blank_row)

    return pd.concat(result_chunks, axis=0)


def create_excel_from_aligned(aligned_dir, output_file):
    texts = load_aligned_texts(aligned_dir)
    if not texts:
        print(f"No aligned files found in {aligned_dir}")
        return

    print(f"Generating Excel from {len(texts)} files...")
    rows = align_to_words(texts)

    data = {}
    for i, t in enumerate(texts):
        data[t["name"]] = rows[i]

    df = pd.DataFrame.from_dict(data, orient="index")

    # Create Excel with two sheets: Original and Printable
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Original sheet - full width
        df.to_excel(writer, sheet_name="Original", header=False)

        # Printable sheet - chunked for A4 landscape
        chunked_df = create_printable_chunks(df, chunk_size=20)
        chunked_df.to_excel(writer, sheet_name="Printable", header=False)

    _apply_excel_formatting(output_file, num_sources=len(df.index), chunk_size=20)

    print(f"Written Excel alignment to {output_file}")
    print(f"  - 'Original' tab: Full alignment ({len(df.columns)} columns)")
    print("  - 'Printable' tab: Chunked for A4 printing (20 columns per chunk)")


def add_printable_tab_to_excel(input_file, output_file, chunk_size=20):
    """
    Reads an existing Excel file, takes the first sheet,
    and appends a new 'Printable' sheet chunked for printing.
    Keeps the existing sheets as they are.
    """
    # Read all sheets to preserve them
    all_sheets = pd.read_excel(input_file, header=None, index_col=0, sheet_name=None)
    if not all_sheets:
        raise ValueError("The provided Excel file has no sheets.")

    first_sheet_name = list(all_sheets.keys())[0]
    first_df = all_sheets[first_sheet_name]

    # Create Excel with original sheets and new Printable sheet
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Write existing sheets
        for sheet_name, sheet_df in all_sheets.items():
            # If the user uploads an excel that already has Printable, skip the old one
            if sheet_name == "Printable":
                continue
            sheet_df.to_excel(writer, sheet_name=sheet_name, header=False)

        # Printable sheet - chunked for A4 landscape based on the first sheet
        chunked_df = create_printable_chunks(first_df, chunk_size=chunk_size)
        chunked_df.to_excel(writer, sheet_name="Printable", header=False)

    _apply_excel_formatting(output_file, num_sources=len(first_df.index), chunk_size=chunk_size)

    print(f"Added 'Printable' tab to {output_file}")

def _apply_excel_formatting(output_file, num_sources, chunk_size):
    wb = openpyxl.load_workbook(output_file)
    thick = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.sheet_view.rightToLeft = True

        bold_font = Font(bold=True)
        for cell in ws["A"]:
            cell.font = bold_font

        # Auto-resize columns
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    val_str = str(cell.value) if cell.value is not None else ""
                    if len(val_str) > max_length:
                        max_length = len(val_str)
                except Exception:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Apply borders for Printable chunks array
        if sheet_name == "Printable":
            max_row = ws.max_row
            total_cols = chunk_size + 1  # Data cols + Column A index
            
            # Form chunks spaced by a single blank delimiter
            for start_row in range(1, max_row + 1, num_sources + 1):
                end_row = min(start_row + num_sources - 1, max_row)
                
                # Check target chunk bounds don't hit the empty filler row exclusively
                if not ws.cell(row=start_row, column=1).value:
                    continue  
                
                # Paint the border perimeter
                for r in range(start_row, end_row + 1):
                    for c in range(1, total_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        borders = {'top': thin, 'bottom': thin, 'left': thin, 'right': thin}
                        
                        if r == start_row:
                            borders['top'] = thick
                        if r == end_row:
                            borders['bottom'] = thick
                        if c == 1:
                            borders['left'] = thick
                        if c == total_cols:
                            borders['right'] = thick
                            
                        cell.border = Border(**borders)

    wb.save(output_file)


def main():
    create_excel_from_aligned(".", "alignment_table.xlsx")


if __name__ == "__main__":
    main()
